# coding: utf8
import os
os.environ['TIKA_SERVER_JAR'] = r'file:///C|/Users/TP2107034\Desktop/ner-spacy-doccano/tika-server-1.24.jar'
os.environ['TIKA_STARTUP_SLEEP'] = '0.2'

from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tika.tika import TikaClientOnly
from ttkthemes import ThemedStyle
from PIL import Image, ImageTk
from pdf2image import convert_from_path
import pandas as pd
from typing import Type
import os
from os import linesep
import plac
import spacy
import re
from ner_train import test_model
import tika
from tika import tika as tk
from tika import parser
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import PurePath, PurePosixPath
import ssl
# tika.initVM()
class JDparser:
    def __init__(self, master: Tk):
        tika.TikaClientOnly = True
        tk.startServer(os.getenv('TIKA_SERVER_JAR'))
        os.environ["PYTHONIOENCODING"]="utf8"
        self.master = master
        self.REQUIRE = ""
        self.HEADER = [
            "Experience Required", "Preferred Experience", "Keys to Success in this Role", "Quialifications", "What we need to see", "Ways to stand out from the crowd", "Plus","Key Responsibilites","Requirements","Education"
        ]
        self.model_path = "./model"
        self.master.title("JDparser")
        self.excel_rowtext = []
        # set style of root
        style = ThemedStyle(self.master)
        style.set_theme("breeze")
        # fixed size
        self.master.geometry("1100x500")
        self.master.resizable(0, 0)
        # Menu
        self.my_menu = Menu(self.master)
        self.master.config(menu=self.my_menu)
        file_menu = Menu(self.my_menu, tearoff=False)
        self.my_menu.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open", command=self.show_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.master.quit)
        # pdf
        self.pdf_frame = Frame(self.master, width=500, height=460)
        self.pdf_frame.pack(side=LEFT, fill=BOTH, expand=0)
        self.fileopening = False
        # adding scrollbar to the PDF frame
        self.scrol_y_pdf = Scrollbar(self.pdf_frame, orient=VERTICAL)
        self.scrol_x_pdf = Scrollbar(self.pdf_frame, orient=HORIZONTAL)
        self.pdf = Text(self.pdf_frame, yscrollcommand=self.scrol_y_pdf.set,
                        xscrollcommand=self.scrol_x_pdf.set, bg='grey')
        # Setting the scrollbar to the right side
        self.scrol_y_pdf.pack(side=RIGHT, fill=Y)
        self.scrol_x_pdf.pack(side=BOTTOM, fill=X)
        self.scrol_y_pdf.config(command=self.pdf.yview)
        self.scrol_x_pdf.config(command=self.pdf.xview)
        self.pdf.pack(fill=BOTH, expand=1)
        # bind enter key with insert header to listbox
        self.master.bind("<Return>", self.insert_header)
        self.master.bind("<Delete>", self.delete_header)
        # Button, fixed position
        Button(self.master, text="Delete Header",
               command=self.delete_header).place(x=800, y=10, height=20)
        Button(self.master, text="Select", command=self.select_required_header).place(
            x=900, y=350, height=20, width=100)
        Button(self.master, text="Export as", command=self.save).place(
            x=750, y=425, height=50, width=100)
        # ListBox
        self.Lb1 = Listbox(self.master, bg='white')
        self.Lb1.place(x=600, y=40, width=400, height=300)
        # Insert common header
        print(self.HEADER)
        self.HEADER.sort()
        
        for header in self.HEADER:
            print(header)
            self.Lb1.insert(END, header)
            print(self.Lb1.get(END))
        scrol_y_lb = Scrollbar(self.Lb1)
        scrol_y_lb.pack(side=RIGHT, fill=BOTH)
        self.Lb1.config(yscrollcommand=scrol_y_lb.set)
        scrol_y_lb.config(command=self.Lb1.yview)
        self.Lb2 = Listbox(self.master, bg='white')
        self.Lb2.place(x=600, y=350, width=300, height=20)
        # header entry and labels
        L1 = Label(self.master, text="Header")
        L1.place(x=600, y=10)
        L2 = Label(self.master, text="Required Header: The first section that you may want "
                   "to append in table.\n Every statements after this header will be included.")
        L2.place(x=600, y=370)
        self.header = Entry(self.master, bg='white')
        self.header.place(x=650, y=10)

    def op(self):
        self.fp_in = filedialog.askopenfilename(
            title='選擇', filetypes=[('pdf', '*.pdf'), ('All Files', '*')])
        print(os.path.exists(self.fp_in))
        return self.fp_in

    def read(self, fp_in):
        self.df = pd.read_excel(fp_in, header=0)
        cols = list(self.df.columns)
        return self.df, cols

    def show(self, frame, df, cols):
        tree = ttk.Treeview(frame)
        tree.pack()
        tree["columns"] = cols
        for i in cols:
            tree.column(i, anchor="center")
            tree.heading(i, text=i, anchor='center')
        for index, row in df.iterrows():
            tree.insert("", 'end', text=index, values=list(row))
        return tree

    def oas(self):
        global root
        self.op()
        data, c = self.read(self.fp_in)
        tree = self.show(self.master, data, c)
        tree.place(relx=0, rely=0.1, relheight=0.7, relwidth=1)

    def save(self):
        self.fp_out = filedialog.asksaveasfilename(filetypes=(("Excel files", "*.xlsx"),
                                                              ("All files", "*.*")))
        # store workbook
        self.parse()

        # note: this will fail unless user ends the fname with ".xlsx"
        print(self.fp_out[-5:])
        if(self.fp_out[-5:] != ".xlsx"):
            self.fp_out += ".xlsx"
        self.wb.save(f"{self.fp_out}")
        print(f"{self.fp_out} is stored.")

    def insert_header(self, event=None):
        text = self.header.get()
        self.Lb1.insert("end", text)
        self.HEADER.append(text)
        self.header.delete(0, END)

    def delete_header(self, event=None):
        index = self.Lb1.curselection()
        self.Lb1.delete(index)

    def select_required_header(self):
        self.Lb2.delete(0, END)
        index = self.Lb1.curselection()
        required_header = self.Lb1.get(index)
        self.Lb2.insert("end", required_header)
        self.REQUIRE = required_header
        self.REQUIRE = re.sub(' +', ' ', self.REQUIRE)
        self.REQUIRE = re.sub('\t', '', self.REQUIRE)

    def show_pdf(self):
        if self.fileopening:
           self.pdf.delete(1.0,END)
            # for widget in self.pdf_frame.winfo_children():
            #     widget.destroy()
        self.fileopening=True
        self.op()
        # Here the PDF is converted to list of images
        pages = convert_from_path(self.fp_in, size=(800, 900))
        # Empty list for storing images
        photos = []
        # Storing the converted images into list
        for i in range(len(pages)):
            photos.append(ImageTk.PhotoImage(pages[i]))
            # Adding all the images to the text widget
        for photo in photos:
            self.pdf.image_create(END, image=photo)
        
        # For Seperating the pages
        # pdf.insert(END,'\n\n')
        # Ending of mainloop
        self.pre_load()
        self.master.mainloop()

    def pre_load(self):
        self.raw_text = []
            # contents = f.read()
            # Parsed pdf with tika
            ## TODO ##
            # run tika offline
        parsed = parser.from_file(self.fp_in)
        contents = parsed['content']
        # print(contents)
        # Removed default texts in reuslts
        contents = contents.replace('http://node.js/', '')
        # Remove non-breaking space by an utf space
        contents = contents.replace(u'\xa0', ' ')
        # Remove redundant linebreak
        contents = re.sub('\n+', '\n', contents)
        # First split lines by bullitin symbols
        self.lines = re.split("[^A-Za-z0-9\-\#\'\\ /&,.+()\n]", contents)

    def pre_parsing(self):
        

        # Require found or not
        found = False
        self.raw_text.clear()

        for line in self.lines:
            # remove redundant blanks
            line = re.sub(' +', ' ', line)
            # line = re.sub('[^A-Za-z0-9\-\#\'\\ /&,.+()]*',"",line)
            line = re.sub('[\n]*', "", line)
            if found:
                # Pass if it is header
                if any(ele in line for ele in self.HEADER):
                    continue
                # Make sure the text has words in it.
                if re.search("[A-Za-z0-9]+", line):
                    self.raw_text.append(line)

            # REQUIRE header found, start appending
            if line.find(str(self.REQUIRE)):
                found = True

    def table_construction(self):
        self.wb = load_workbook(filename="templates.xlsx")
        
        template = self.wb['templates']
        template_h = 8
        template.column_dimensions["C"].auto_size = True
        
        # fontStyle = Font(size = "18")
        double = Side(border_style="thin")
        dataval = DataValidation(
            type="list", formula1='"No Experience(0),Basic(1-2),Intermediated(3-4),Advanced(5)"')
        template.add_data_validation(dataval)
        print("len:" + str(len(self.excel_rowtext)))
        for r in range(len(self.excel_rowtext)+10):
            print(r)
            template.row_dimensions[r+template_h].height = 45
            for c in range(len(self.excel_rowtext[0])-1):
                template.cell(row=r+template_h, column=2+c).border = Border(top=double,
                                                                   left=double, right=double, bottom=double)

                if r < len(self.excel_rowtext):                            # statement
                    if self.excel_rowtext[r][5] == '1':
                        
                        template.cell(
                            row=r+template_h, column=2+c).font = Font(name='Tahoma',size=14,  bold=True)
                        template.cell(
                            row=r+template_h, column=2+c).fill = PatternFill("solid", fgColor='CCFFFF')
                    else:
                        template.cell(row=r+template_h, column=2 +
                                      c).font = Font(name='Tahoma',size=14)
                    template.cell(row=r+template_h, column=2 +
                                  c).value = self.excel_rowtext[r][c]
                else:
                    template.cell(row=r+template_h, column=2+c).value = ""
                template.cell(row=r+template_h, column=2+c).alignment =Alignment(wrapText=True)
                dataval.add(template.cell(row=r+template_h, column=5))
                template.row_dimensions[r+template_h].height = 50
                

        # wb.save(f"{PurePosixPath(path).stem}.xlsx")
        # print(f"{PurePosixPath(path).stem}.xlsx is saved.")

    def parse(self):
        self.nlp = spacy.load(self.model_path)
        self.pre_parsing()
        # This list contains structured format of statements
        # [No.,Text ,Year, proficiency level, discription]
        if len(self.excel_rowtext) > 0:
            print("CLEAR")
            self.excel_rowtext.clear()
        
        # first column of excel
        num = 1
        for item in self.raw_text:
            doc = self.nlp(item)
            yr = ""
            for entity in doc.ents:
                ## TODO##
                # How to solve multiple years in one statement
                # only store first year
                if entity.label_ == "YEAR":
                    yr = entity.text
                    break

            a1 = [num, item, yr, '', '', 1]

            self.excel_rowtext.append(a1)
            num += 1
            for entity in doc.ents:
                if entity.label_ == 'SKILL':
                    a2 = ['', entity.text, '', '', '', 0]
                    self.excel_rowtext.append(a2)
        print(self.raw_text)
        self.table_construction()


if __name__ == '__main__':
    root = Tk()
    jdparser = JDparser(root)
    root.mainloop()
