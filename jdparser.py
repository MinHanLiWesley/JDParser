# coding: utf8
import os
os.environ['TIKA_SERVER_JAR'] = r'file:///C|/Users/TP2107034\Desktop/ner-spacy-doccano/tika-server-1.24.jar'
from pathlib import PurePath, PurePosixPath
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from tika import parser
from os import linesep
import plac
import spacy
import re

from tika.tika import TikaVersion
from ner_train import test_model
import tika
# tika.initVM()


## TODO ##
# Add more header or create an interface to input
HEADER = [
    'Experience‌ ‌Required‌', 'Preferred‌ ‌Experience‌', 'Keys‌ ‌to‌ ‌Success‌ ‌in‌ ‌this‌ ‌Role', 'Qualifications', 'What we need to see', 'Ways to stand out from the crowd', 'Plus', 'Keys to Success in this Role'
]
REQUIRE = ['require', 'qualification', 'Require', 'Qualification']

PATH = r"C:\Users\TP2107034\Desktop\ner-spacy-doccano\Mobile_App_Developer_JD_-_Google_Docs.pdf"


def load_data(path):
    # os.environ['TIKA_SERVER_JAR'] = "file:///c:/Users/TP2107034/Desktop/ner-spacy-doccano/tika-server-standard-2.1.0.jar"
    os.environ['TIKA_SERVER_JAR'] = 'https://repo1.maven.org/maven2/org/apache/tika/tika-server/1.19/tika-server-1.19.jar'
    TikaVersion='1.24'
    print(os.getenv(
    'TIKA_SERVER_JAR',
    "http://search.maven.org/remotecontent?filepath=org/apache/tika/tika-server/"+TikaVersion+"/tika-server-"+TikaVersion+".jar"))
    print(os.getenv('TIKA_SERVER_JAR'))

    raw_text=[]
    with open(path,'r',encoding='unicode_escape') as f:
        contents = f.read()
        # Parsed pdf with tika
        ## TODO ##
        # run tika offline
        parsed = parser.from_file(path)
        contents = parsed['content']
        # Removed default texts in reuslts
        contents = contents.replace('http://node.js/','')
        # Remove non-breaking space by an utf space
        contents = contents.replace(u'\xa0',' ')
        # Remove redundant linebreak
        contents = re.sub('\n+', '\n', contents)
        # First split lines by bullitin symbols
        lines = re.split("[^A-Za-z0-9\-\#\'\\ /&,.+()\n]",contents)

        # Require found or not
        found = False

        for line in lines:
            # remove redundant blanks
            line = re.sub(' +', ' ', line)
            # line = re.sub('[^A-Za-z0-9\-\#\'\\ /&,.+()]*',"",line)
            line = re.sub('[\n]*',"",line)
            if found:
                # Pass if it is header
                if any(ele in line for ele in HEADER):
                    continue
                # Make sure the text has words in it.
                if re.search("[A-Za-z0-9]+",line):
                    raw_text.append(line)
                
            # REQUIRE header found, start appending
            if any(ele in line for ele in REQUIRE):
                found = True

    return raw_text

def table_construction(text2,path):
    wb = load_workbook(filename="templates.xlsx")
    template = wb['templates']
    double = Side(border_style="thin")
    dataval = DataValidation(type="list", formula1='"No Experience(0),Basic(1-2),Intermediated(3-4),Advanced(5)"')
    template.add_data_validation(dataval)
    for r in range(len(text2)+10):
        for c in range(len(text2[0])-1):
            template.cell(row=9+r,column=2+c).border = Border(top=double, left=double, right=double, bottom=double)

            if r < len(text2):                            # statement
                if text2[r][5] == '1':
                    template.cell(row=9+r,column=2+c).font = Font(name='Tahoma',size=12,bold=True)
                    template.cell(row=9+r,column=2+c).fill = PatternFill("solid", fgColor = 'CCFFFF')
                else:
                    template.cell(row=9+r,column=2+c).font = Font(name='Tahoma',size=12)
                template.cell(row=9+r,column=2+c).value = text2[r][c]
            else:
                template.cell(row=9+r,column=2+c).value = ""
            dataval.add(template.cell(row=9+r,column=5)) 
                
             
    wb.save(f"{PurePosixPath(path).stem}.xlsx")
    print(f"{PurePosixPath(path).stem}.xlsx is saved.")
    


@plac.annotations(
    model=("Model_name.", "option", "m", str),
    path=("pdf_path.", "option", 'p',str)
)
def main(model='./model',path=PATH):
    print("Loading from", model)
    nlp = spacy.load(model)

    DATA = load_data(path)
    # test_model(nlp, DATA)
    # This list contains structured format of statements
    # [No.,Text ,Year, proficiency level, discription]
    excel_rowtext=[]
    # first column of excel
    num=1
    for item in DATA:
        doc = nlp(item)
        yr=""
        for entity in doc.ents:
            ## TODO## 
            # How to solve multiple years in one statement
            # only store first year
            if entity.label_ == "YEAR":
                yr = entity.text
                break

        a1 = [num,item,yr,'','',1]

        excel_rowtext.append(a1)
        num+=1
        for entity in doc.ents:
            if entity.label_ =='SKILL':
                a2 = ['',entity.text,'','','',0]
                excel_rowtext.append(a2)
    table_construction(excel_rowtext,path)
    


if __name__ == "__main__":
    plac.call(main)
